# ruff: noqa: INP001, T201, E501
"""Generate metadata feedback report comparing raw vs corrected metadata.

Outputs a markdown summary (.md) and an Excel workbook (.xlsx) with:
  - Sheet "Raw"       — raw source data with changed cells highlighted
  - Sheet "Corrected" — corrected metadata_all with changed cells highlighted

Usage:
    uv run python scripts/generate_metadata_feedback.py [data_dir]

Output files are written to ./tmp/.
"""

from __future__ import annotations

import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from hdx.scraper.cod_ab_country.download.metadata.process import (
    column_rename,
)

MERGE_KEYS: list[str] = ["country_iso3", "version"]

YELLOW_FILL: PatternFill = PatternFill("solid", fgColor="FFEB9C")  # Neutral
GREEN_FILL: PatternFill = PatternFill("solid", fgColor="C6EFCE")  # Good
ORANGE_FILL: PatternFill = PatternFill("solid", fgColor="FFCC99")  # Input
REMOVED_FILL: PatternFill = PatternFill("solid", fgColor="FFC7CE")  # Bad
COERCE_FILL: PatternFill = PatternFill("solid", fgColor="BDD7EE")  # Note (text→number)
HEADER_FONT: Font = Font(bold=True)

_REVERSED_RENAME: dict[str, str] = {v: k for k, v in column_rename.items()}


def _is_na(val: object) -> bool:
    """Return True if val is a missing/null value."""
    if val is None:
        return True
    try:
        return bool(pd.isna(val))  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return False


def _str_val(val: object) -> str:
    """Normalize val to a stripped string for comparison."""
    return "" if _is_na(val) else str(val).strip()


def _vals_equal(a: object, b: object) -> bool:
    """Return True if two values are semantically equal."""
    if _is_na(a) and _is_na(b):
        return True
    if _is_na(a) or _is_na(b):
        return False
    sa, sb = _str_val(a), _str_val(b)
    if sa == sb:
        return True
    try:
        return float(sa) == float(sb)
    except ValueError:
        return False


def _is_type_coercion(a: object, b: object) -> bool:
    """Return True if a is a string that represents the same number as b (numeric)."""
    if _is_na(a) or _is_na(b):
        return False
    if not isinstance(a, str) or isinstance(b, str):
        return False
    try:
        return float(_str_val(a)) == float(_str_val(b))
    except ValueError:
        return False


def _infer_change_type(field: str, raw_val: object) -> str:
    """Infer the category of a cell-level change."""
    if field == "source":
        return "source_correction"
    if field == "contributor":
        return "contributor_correction"
    if field == "admin_level_full":
        return "admin_level_correction"
    if _str_val(raw_val).lower() == "currently not known":
        return "null_cleanup"
    if field in {"country_iso2", "country_name"}:
        return "derived_field"
    return "other"


def _to_cell_value(val: object) -> str | int | float | None:
    """Convert a pandas/numpy value to an Excel-safe Python type."""
    if _is_na(val):
        return None
    if hasattr(val, "item"):
        return val.item()  # type: ignore[union-attr]
    if isinstance(val, pd.Timestamp):
        return val.date().isoformat()
    if isinstance(val, datetime):
        return val.date().isoformat()
    return val  # type: ignore[return-value]


def build_schema_migrations(
    df_raw: pd.DataFrame,
    df_all: pd.DataFrame,
) -> list[dict[str, str]]:
    """Return columns where dtype differs between raw and processed tables."""
    df_raw_r = df_raw.rename(columns=column_rename)
    return [
        {
            "field": col,
            "raw_dtype": str(df_raw_r[col].dtype),
            "corrected_dtype": str(df_all[col].dtype),
        }
        for col in df_all.columns
        if col in df_raw_r.columns
        and str(df_raw_r[col].dtype) != str(df_all[col].dtype)
    ]


def build_changes(
    df_raw: pd.DataFrame,
    df_all: pd.DataFrame,
) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    """Compare raw and all metadata tables.

    Returns (cell_changes, added_rows, removed_rows) where row IDs are
    'ISO3 version' strings.
    """
    df_raw_r = df_raw.rename(columns=column_rename)
    all_cols = set(df_all.columns) - set(MERGE_KEYS)
    compare_cols = [
        c for c in df_all.columns if c in all_cols and c in df_raw_r.columns
    ]

    raw_idx = df_raw_r.set_index(MERGE_KEYS)
    all_idx = df_all.set_index(MERGE_KEYS)

    changes: list[dict[str, Any]] = []
    added_rows: list[str] = []

    for idx in all_idx.index:
        all_row = all_idx.loc[idx]
        if isinstance(all_row, pd.DataFrame):
            all_row = all_row.iloc[0]

        if idx in raw_idx.index:
            raw_row = raw_idx.loc[idx]
            if isinstance(raw_row, pd.DataFrame):
                raw_row = raw_row.iloc[0]

            for col in compare_cols:
                raw_val = raw_row[col] if col in raw_row.index else None
                all_val = all_row[col] if col in all_row.index else None
                is_coercion = _is_type_coercion(raw_val, all_val)
                if not _vals_equal(raw_val, all_val) or is_coercion:
                    change_type = (
                        "type_coercion"
                        if is_coercion
                        else _infer_change_type(col, raw_val)
                    )
                    changes.append(
                        {
                            "country_iso3": idx[0],
                            "version": idx[1],
                            "field": col,
                            "raw_value": _to_cell_value(raw_val),
                            "corrected_value": _to_cell_value(all_val),
                            "change_type": change_type,
                        }
                    )
        else:
            added_rows.append(f"{idx[0]} {idx[1]}")

    removed_rows = [
        f"{idx[0]} {idx[1]}" for idx in raw_idx.index if idx not in all_idx.index
    ]

    return changes, added_rows, removed_rows


def _md_table(headers: list[str], rows: list[list[str]]) -> list[str]:
    """Render a markdown table as a list of lines."""
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    return [
        "| " + " | ".join(headers) + " |",
        sep,
        *("| " + " | ".join(r) + " |" for r in rows),
    ]


def _md_val(val: object) -> str:
    """Format a value for a markdown table cell."""
    if val is None or _str_val(val) == "":
        return "_(blank)_"
    return str(val).replace("|", "\\|")


def build_markdown_summary(
    changes: list[dict[str, Any]],
    added_rows: list[str],
    removed_rows: list[str],
    schema_migrations: list[dict[str, str]] | None = None,
) -> str:
    """Build a markdown summary of all metadata corrections."""
    today = datetime.now(tz=UTC).date().isoformat()
    lines: list[str] = ["# Metadata Feedback Summary", "", f"Generated: {today}", ""]

    if schema_migrations:
        lines += [
            f"## Schema Migrations — {len(schema_migrations)} columns",
            "",
            "These columns have a different dtype in the source table than in our output:",
            "",
            *_md_table(
                ["Field", "Source dtype", "Output dtype"],
                [
                    [
                        f"`{m['field']}`",
                        f"`{m['raw_dtype']}`",
                        f"`{m['corrected_dtype']}`",
                    ]
                    for m in schema_migrations
                ],
            ),
            "",
        ]

    lines += [
        f"## Column Renames — {len(column_rename)}",
        "",
        "These field names in the source table contain typos:",
        "",
        *_md_table(
            ["Source name", "Correct name"],
            [[f"`{old}`", f"`{new}`"] for old, new in column_rename.items()],
        ),
        "",
    ]

    src = [c for c in changes if c["change_type"] == "source_correction"]
    lines += [f"## Source Corrections — {len(src)} entries", ""]
    lines += _md_table(
        ["Country", "Version", "Current value", "Suggested value"],
        [
            [
                c["country_iso3"],
                c["version"],
                _md_val(c["raw_value"]),
                _md_val(c["corrected_value"]),
            ]
            for c in sorted(src, key=lambda x: x["country_iso3"])
        ],
    )
    lines.append("")

    contrib = [c for c in changes if c["change_type"] == "contributor_correction"]
    lines += [f"## Contributor Corrections — {len(contrib)} entries", ""]
    lines += _md_table(
        ["Country", "Version", "Current value", "Suggested value"],
        [
            [
                c["country_iso3"],
                c["version"],
                _md_val(c["raw_value"]),
                _md_val(c["corrected_value"]),
            ]
            for c in sorted(contrib, key=lambda x: x["country_iso3"])
        ],
    )
    lines.append("")

    admin = [c for c in changes if c["change_type"] == "admin_level_correction"]
    lines += [f"## `admin_level_full` Corrections — {len(admin)} entries", ""]
    lines += _md_table(
        ["Country", "Version", "Current value", "Suggested value"],
        [
            [
                c["country_iso3"],
                c["version"],
                _md_val(c["raw_value"]),
                _md_val(c["corrected_value"]),
            ]
            for c in sorted(admin, key=lambda x: (x["country_iso3"], x["version"]))
        ],
    )
    lines.append("")

    null_changes = [c for c in changes if c["change_type"] == "null_cleanup"]
    countries_with_null = sorted({c["country_iso3"] for c in null_changes})
    lines += [
        f"## Null Cleanup"
        f" — {len(null_changes)} cells across {len(countries_with_null)} countries",
        "",
        'The string `"currently not known"` should be blank/null in these fields:',
        "",
        *_md_table(
            ["Country", "Fields"],
            [
                [
                    iso3,
                    ", ".join(
                        f"`{f}`"
                        for f in sorted(
                            {
                                c["field"]
                                for c in null_changes
                                if c["country_iso3"] == iso3
                            }
                        )
                    ),
                ]
                for iso3 in countries_with_null
            ],
        ),
        "",
    ]

    derived = [c for c in changes if c["change_type"] == "derived_field"]
    lines += [
        f"## Derived Fields — {len(derived)} cells",
        "",
        "These fields are blank in the source table but populated by us from the ISO3 code:",
        "",
    ]
    for field in sorted({c["field"] for c in derived}):
        count = sum(1 for c in derived if c["field"] == field)
        lines.append(f"- `{field}`: {count} rows")
    lines.append("")

    lines += [f"## Missing Rows — {len(added_rows)}", ""]
    if added_rows:
        lines.append("These rows are absent from the source table and should be added:")
        lines.append("")
        lines += _md_table(
            ["Country", "Version"],
            [r.split(" ", 1) for r in sorted(added_rows)],
        )
    lines.append("")

    lines += [f"## Rows Excluded from Our Output — {len(removed_rows)}", ""]
    if removed_rows:
        lines.append(
            "These rows are in the source table but excluded from our published output:"
        )
        lines.append("")
        lines.append(", ".join(sorted(removed_rows)))
    lines.append("")

    other = [c for c in changes if c["change_type"] == "other"]
    if other:
        key = lambda x: (x["country_iso3"], x["version"], x["field"])  # noqa: E731
        lines += [f"## Other Changes — {len(other)}", ""]
        lines += _md_table(
            ["Country", "Version", "Field", "Current value", "Suggested value"],
            [
                [
                    c["country_iso3"],
                    c["version"],
                    f"`{c['field']}`",
                    _md_val(c["raw_value"]),
                    _md_val(c["corrected_value"]),
                ]
                for c in sorted(other, key=key)
            ],
        )
        lines.append("")

    return "\n".join(lines)


def _auto_width(ws: Any) -> None:  # noqa: ANN401
    """Set column widths based on content."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _write_raw_sheet(
    wb: Workbook,
    df_raw: pd.DataFrame,
    changes: list[dict[str, Any]],
    removed_rows: list[str],
) -> None:
    """Write the Raw sheet — original source data with changed cells highlighted."""
    ws = wb.active
    ws.title = "Raw"  # type: ignore[union-attr]

    # Build highlight lookup using raw column names (reverse the rename mapping)
    changed_cells_raw: set[tuple[str, str, str]] = {
        (c["country_iso3"], c["version"], _REVERSED_RENAME.get(c["field"], c["field"]))
        for c in changes
        if c["change_type"] != "type_coercion"
    }
    coerced_cells_raw: set[tuple[str, str, str]] = {
        (c["country_iso3"], c["version"], _REVERSED_RENAME.get(c["field"], c["field"]))
        for c in changes
        if c["change_type"] == "type_coercion"
    }
    removed_set: set[tuple[str, str]] = {
        (r.split(" ", 1)[0], r.split(" ", 1)[1]) for r in removed_rows
    }

    headers = list(df_raw.columns)
    ws.append(headers)  # type: ignore[union-attr]
    for cell in ws[1]:  # type: ignore[union-attr]
        cell.font = HEADER_FONT
        if cell.value in column_rename:
            cell.fill = ORANGE_FILL
            cell.comment = Comment(
                f"Rename to: {column_rename[cell.value]}", "feedback-script"
            )

    for _, row in df_raw.iterrows():
        iso3 = row["country_iso3"]
        ver = row["version"]
        ws.append([_to_cell_value(row[col]) for col in headers])  # type: ignore[union-attr]
        row_idx = ws.max_row  # type: ignore[union-attr]
        is_removed = (iso3, ver) in removed_set
        for col_idx, col in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore[union-attr]
            if is_removed:
                cell.fill = REMOVED_FILL
            elif (iso3, ver, col) in coerced_cells_raw:
                cell.fill = COERCE_FILL
            elif (iso3, ver, col) in changed_cells_raw:
                cell.fill = YELLOW_FILL

    _auto_width(ws)
    ws.freeze_panes = "A2"  # type: ignore[union-attr]


def _write_corrected_sheet(  # noqa: C901, PLR0912, PLR0913
    wb: Workbook,
    df_raw: pd.DataFrame,
    df_all: pd.DataFrame,
    changes: list[dict[str, Any]],
    added_rows: list[str],
    removed_rows: list[str],
) -> None:
    """Write the Corrected sheet — raw data with only tracked changes applied."""
    ws = wb.create_sheet("Corrected")

    # Build lookup of corrected values keyed by (iso3, version, raw_col)
    change_lookup: dict[tuple[str, str, str], object] = {}
    coercion_keys: set[tuple[str, str, str]] = set()
    for c in changes:
        raw_col = _REVERSED_RENAME.get(c["field"], c["field"])
        key = (c["country_iso3"], c["version"], raw_col)
        change_lookup[key] = c["corrected_value"]
        if c["change_type"] == "type_coercion":
            coercion_keys.add(key)

    removed_set: set[tuple[str, str]] = {
        (r.split(" ", 1)[0], r.split(" ", 1)[1]) for r in removed_rows
    }
    added_set: set[tuple[str, str]] = {
        (r.split(" ", 1)[0], r.split(" ", 1)[1]) for r in added_rows
    }

    headers = list(df_raw.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        if cell.value in column_rename:
            cell.fill = ORANGE_FILL
            cell.comment = Comment(
                f"Rename to: {column_rename[cell.value]}", "feedback-script"
            )

    for _, row in df_raw.iterrows():
        iso3 = str(row["country_iso3"])
        ver = str(row["version"])
        if (iso3, ver) in removed_set:
            continue
        out_row = []
        for col in headers:
            key = (iso3, ver, col)
            out_row.append(
                change_lookup[key] if key in change_lookup else _to_cell_value(row[col])
            )
        ws.append(out_row)
        row_idx = ws.max_row
        for col_idx, col in enumerate(headers, start=1):
            key = (iso3, ver, col)
            cell = ws.cell(row=row_idx, column=col_idx)
            if key in coercion_keys:
                cell.fill = COERCE_FILL
            elif key in change_lookup:
                cell.fill = YELLOW_FILL

    # Append rows that exist in corrected output but not in raw
    if added_set:
        df_added = df_all.rename(columns=_REVERSED_RENAME).reindex(columns=headers)
        mask = df_added.apply(
            lambda r: (str(r["country_iso3"]), str(r["version"])) in added_set, axis=1
        )
        for _, row in df_added[mask].iterrows():
            ws.append([_to_cell_value(row[col]) for col in headers])
            row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = GREEN_FILL

    _auto_width(ws)
    ws.freeze_panes = "A2"


def main() -> None:
    """Entry point."""
    data_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("saved_data/metadata")
    out_dir = Path("tmp")
    out_dir.mkdir(exist_ok=True)

    df_raw = pd.read_parquet(data_dir / "metadata_raw.parquet")
    df_all = pd.read_parquet(data_dir / "metadata_all.parquet")

    changes, added_rows, removed_rows = build_changes(df_raw, df_all)
    schema_migrations = build_schema_migrations(df_raw, df_all)
    md = build_markdown_summary(changes, added_rows, removed_rows, schema_migrations)
    print(md)

    today = datetime.now(tz=UTC).date().isoformat()
    stem = f"feedback_metadata_{today}"

    md_path = out_dir / f"{stem}.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"Markdown report saved: {md_path}")

    wb = Workbook()
    _write_raw_sheet(wb, df_raw, changes, removed_rows)
    _write_corrected_sheet(wb, df_raw, df_all, changes, added_rows, removed_rows)

    xlsx_path = out_dir / f"{stem}.xlsx"
    wb.save(xlsx_path)
    print(f"Excel report saved: {xlsx_path}")


if __name__ == "__main__":
    main()
